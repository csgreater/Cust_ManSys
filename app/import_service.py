from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook


HEADER_MAP = {
    "商品链接Id": "link_id",
    "商品链接SkuId": "sku_id",
    "订单来源": "order_source",
    "客户编号": "customer_no",
    "客户名称": "customer_name",
    "渠道分类": "dept",
    "渠道平台": "platform",
    "销售渠道": "shop_name",
    "订单编号": "order_no",
    "网店订单号（新）": "original_order_no",
    "物流公司": "logistics_type",
    "物流单号": "logistics_no",
    "收货人": "receiver_name",
    "地址": "receiver_address",
    "电话": "receiver_phone",
    "大类": "category",
    "货品分类": "product_classification",
    "货品名称": "product_name",
    "货品编号": "product_no",
    "单位": "unit",
    "数量": "qty",
    "销售额": "share_receivable",
    "省": "province",
    "市": "city",
    "县": "district",
    "发货时间": "ship_time",
    "成本金额": "cost",
    "快递费": "express_fee",
    "物流费": "logistics_fee",
    "运费": "freight",
    "辅料费用": "aux_material",
    "分摊费用": "share_cost",
    "利润": "excel_profit",
}

OPTIONAL_HEADERS = {"货品分类"}
REQUIRED_HEADERS = [header for header in HEADER_MAP if header not in OPTIONAL_HEADERS]
REQUIRED_FIELDS = {
    "order_no",
    "customer_no",
    "dept",
    "platform",
    "shop_name",
    "product_no",
    "ship_time",
    "share_receivable",
    "cost",
    "excel_profit",
}
DECIMAL_FIELDS = {
    "qty",
    "share_receivable",
    "cost",
    "express_fee",
    "logistics_fee",
    "freight",
    "aux_material",
    "share_cost",
    "excel_profit",
}
DECIMAL_MAX_ABS = {
    "qty": Decimal("99999999.99"),
    **{
        field: Decimal("9999999999.99")
        for field in DECIMAL_FIELDS
        if field != "qty"
    },
}
NON_NEGATIVE_FIELDS: set[str] = set()
DEFAULT_SHIP_TIME = datetime(1970, 1, 1)
TEXT_MAX_LENGTHS = {
    "link_id": 64,
    "sku_id": 64,
    "order_source": 32,
    "customer_no": 32,
    "customer_name": 64,
    "dept": 32,
    "platform": 32,
    "shop_name": 64,
    "order_no": 128,
    "original_order_no": 512,
    "logistics_type": 32,
    "logistics_no": 64,
    "receiver_name": 32,
    "receiver_address": 512,
    "receiver_phone": 32,
    "category": 32,
    "product_classification": 64,
    "product_name": 255,
    "product_no": 64,
    "unit": 16,
    "province": 32,
    "city": 64,
    "district": 64,
}


@dataclass(frozen=True)
class ImportParseResult:
    batch_no: str
    rows: list[dict[str, Any]]
    total_rows: int
    fail_rows: int
    error_summary: str


def new_batch_no() -> str:
    return f"IMP{datetime.now():%Y%m%d%H%M%S}_{uuid4().hex[:8].upper()}"


def to_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        raise ValueError("不是有效数字")
    if not result.is_finite():
        raise ValueError("必须是有限数字")
    return result


def to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(field: str, value: Any) -> str:
    text = to_text(value)
    max_length = TEXT_MAX_LENGTHS.get(field)
    if max_length is not None:
        return text[:max_length]
    return text


def to_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError("不是有效日期")


def calc_profit(row: dict[str, Any]) -> Decimal:
    return (
        row["share_receivable"]
        - row["cost"]
        - row["freight"]
        - row["aux_material"]
        - row["share_cost"]
    )


def header_indexes(headers: list[str]) -> dict[str, int]:
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        detected = "、".join(header for header in headers if header)
        expected = "、".join(REQUIRED_HEADERS)
        raise ValueError(
            f"Excel 表头不匹配，缺少字段：{', '.join(missing)}。"
            f"当前识别到：{detected or '空表头'}。"
            f"请使用模板字段：{expected}"
        )
    return {
        HEADER_MAP[header]: headers.index(header)
        for header in HEADER_MAP
        if header in headers
    }


def iter_excel_rows(path: Path, batch_no: str) -> Iterator[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("Excel 文件为空")
        indexes = header_indexes([to_text(value) for value in header_row])
        for row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            item: dict[str, Any] = {
                "batch_no": batch_no,
                "row_no": row_no,
                "product_classification": "",
            }
            errors: list[str] = []
            warnings: list[str] = []
            missing_required: set[str] = set()
            for field, index in indexes.items():
                value = values[index] if index < len(values) else None
                if field in REQUIRED_FIELDS and value in (None, ""):
                    missing_required.add(field)
                try:
                    if field == "ship_time":
                        item[field] = to_datetime(value)
                    elif field in DECIMAL_FIELDS:
                        item[field] = to_decimal(value)
                        if abs(item[field]) > DECIMAL_MAX_ABS[field]:
                            raise ValueError("超出数据库可保存范围")
                    else:
                        item[field] = normalize_text(field, value)
                except ValueError as exc:
                    item[field] = Decimal("0") if field in DECIMAL_FIELDS else None
                    errors.append(f"{field}{exc}")

            for field in REQUIRED_FIELDS:
                if field in missing_required or item.get(field) in (None, ""):
                    errors.append(f"{field}不能为空")
            for field in NON_NEGATIVE_FIELDS:
                if item.get(field, Decimal("0")) < 0:
                    errors.append(f"{field}不能为负数")
            if item.get("ship_time") is None:
                item["ship_time"] = DEFAULT_SHIP_TIME
                warnings.append("发货时间为空，已按 1970-01-01 保存")

            item["profit"] = calc_profit(item)
            excel_profit = item.get("excel_profit")
            if excel_profit not in (None, "") and abs(item["profit"] - excel_profit) > Decimal("0.05"):
                warnings.append("Excel利润与系统计算利润不一致")

            item["error_message"] = "; ".join(errors)
            item["warning_message"] = "; ".join(warnings)
            yield item
    finally:
        wb.close()


def parse_excel(path: Path, batch_no: str | None = None) -> ImportParseResult:
    batch_no = batch_no or new_batch_no()
    rows = list(iter_excel_rows(path, batch_no))
    fail_rows = sum(1 for row in rows if row["error_message"])
    return ImportParseResult(
        batch_no=batch_no,
        rows=rows,
        total_rows=len(rows),
        fail_rows=fail_rows,
        error_summary=f"共 {len(rows)} 行，异常 {fail_rows} 行",
    )
